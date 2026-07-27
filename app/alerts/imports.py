from __future__ import annotations

import csv
import io

from app.alerts.enums import ImportSource
from app.alerts.parser import parse_bulk_text


_ALIASES = {
    "symbol": "symbol", "sembol": "symbol", "hisse": "symbol", "ticker": "symbol",
    "price": "price", "fiyat": "price", "target": "price", "hedef": "price", "alarm": "price",
    "condition": "condition", "koşul": "condition", "kosul": "condition", "direction": "condition",
    "yön": "condition", "yon": "condition", "repeat": "mode", "tekrar": "mode",
    "interval": "interval", "aralık": "interval", "aralik": "interval", "note": "note", "not": "note",
}


def _rows_to_text(rows: list[list[object]]) -> str:
    if not rows:
        return ""
    headers = [_ALIASES.get(str(value).strip().casefold(), "") for value in rows[0]]
    if not {"symbol", "price", "condition"}.issubset(headers):
        return "\n".join(",".join(str(value or "") for value in row) for row in rows)
    output = []
    for row in rows[1:]:
        record = {headers[index]: row[index] for index in range(min(len(row), len(headers))) if headers[index]}
        output.append(",".join(str(record.get(key, "") or "") for key in ("symbol", "price", "condition", "mode", "interval", "note")))
    return "\n".join(output)


def parse_csv_bytes(content: bytes, *, maximum_rows: int = 250):
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1254"):
        try: text = content.decode(encoding); break
        except UnicodeDecodeError: continue
    if text is None:
        raise ValueError("CSV metin kodlaması okunamadı.")
    sample = text[:4096]
    try: dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
    except csv.Error: dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return parse_bulk_text(_rows_to_text(rows), maximum_rows=maximum_rows, source=ImportSource.CSV)


def parse_xlsx_bytes(content: bytes, *, maximum_rows: int = 250):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX desteği için openpyxl kurulmalı.") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    if len(rows) - 1 > maximum_rows:
        raise ValueError(f"Tek seferde en fazla {maximum_rows} alarm işlenebilir.")
    return parse_bulk_text(_rows_to_text(rows), maximum_rows=maximum_rows, source=ImportSource.XLSX)
